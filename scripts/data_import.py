# -*- coding: utf-8 -*-
"""
作品数据文件导入解析
====================
把用户从抖音/剪映等平台导出的数据文件解析成结构化指标，供「录入作品数据」使用。

支持的文件：
  1. 流量数据（.xlsx）：播放量、点赞量、评论量、分享量、收藏量、完播率、2s跳出率
  2. 内容吸引力分析（.xlsx）：完播率、平均播放时长、2s跳出率、5s完播率、平均播放占比
  3. 观众分析（.xlsx）：涨粉量、涨粉率
  4. 字幕稿（.txt）：口播逐字稿

解析策略：xlsx 里「指标名 → 数值」的常见排布（同行的右邻格 / 同列的下邻格）都能识别，
不依赖固定表头顺序。
"""
import os
import re

import openpyxl

# 中文规范键 → 英文展示键（作品卡/贡献榜读取英文键）
ZH_TO_EN = {
    "播放量": "plays",
    "完播率": "completion",
    "点赞量": "likes",
    "评论量": "comments",
    "分享量": "shares",
    "收藏量": "saves",
    "涨粉量": "followers",
    "涨粉率": "followers_rate",
}

# 别名 → 规范中文键（按顺序匹配，先匹配到的优先）
ALIASES = [
    ("播放量", ["播放量", "播放次数", "播放数", "播放"]),
    ("点赞量", ["点赞量", "点赞数", "点赞"]),
    ("评论量", ["评论量", "评论数", "评论"]),
    ("分享量", ["分享量", "分享数", "转发量", "转发数", "转发"]),
    ("收藏量", ["收藏量", "收藏数", "收藏"]),
    ("完播率", ["完播率", "完播"]),
    ("2秒跳出率", ["2秒跳出率", "2s跳出率", "两秒跳出率", "2秒跳出", "2s跳出"]),
    ("5秒完播率", ["5秒完播率", "5s完播率", "五秒完播率", "5秒完播", "5s完播"]),
    ("平均播放时长", ["平均播放时长", "平均播放时间", "平均观看时长", "人均播放时长", "播放时长"]),
    ("平均播放占比", ["平均播放占比", "平均播放比例", "播放占比"]),
    ("涨粉量", ["涨粉量", "涨粉数", "涨粉", "新增粉丝", "新增粉丝量"]),
    ("涨粉率", ["涨粉率", "涨粉比例"]),
]


def _norm_label(s) -> str:
    """把单元格文字归一化，用于和别名做精确比对。"""
    s = str(s).strip()
    s = re.sub(r"[（(].*?[)）]", "", s)          # 去括号注释/单位
    s = re.sub(r"[:：]\s*$", "", s)              # 去结尾冒号
    s = re.sub(r"[\s\u3000]+", "", s)            # 去空格 / 全角空格
    return s


def _match_label(s):
    """返回单元格文字对应的规范中文键；不匹配返回 None。"""
    n = _norm_label(s)
    if not n:
        return None
    for key, aliases in ALIASES:
        for a in aliases:
            if n == _norm_label(a):
                return key
    return None


def _is_value_like(v) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    return bool(re.search(r"\d", str(v)))


# 百分比类指标：统一换算到「0~100 百分数」刻度（与手动录入、榜单单位一致）
PERCENT_KEYS = {"完播率", "2秒跳出率", "5秒完播率", "平均播放占比", "涨粉率", "留存率"}


def _to_value(v, key=None):
    """把单元格值规范成数字（支持「12000」「35%」「18秒」「1.2w」「0.35」等）。"""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        n = float(v)
        if key in PERCENT_KEYS and 0 < n < 1:
            n *= 100
        return int(n) if n.is_integer() else round(n, 4)
    s = str(v).strip().replace(",", "").replace("，", "")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    n = float(m.group(0))
    if "万" in s or re.search(r"w\b", s, re.IGNORECASE):
        n *= 10000
    if "亿" in s:
        n *= 100000000
    # 百分比类指标：无「%」符号且落在 0~1 之间的小数视为比例，转成百分数
    has_pct = ("%" in s) or ("％" in s)
    if key in PERCENT_KEYS and not has_pct and 0 < n < 1:
        n *= 100
    return int(n) if n.is_integer() else round(n, 4)


def normalize_metrics(metrics: dict) -> dict:
    """把指标字典补全为「中文键 + 英文键」双份，供榜单与作品卡同时使用。"""
    out = {}
    for k, v in (metrics or {}).items():
        if v is None or v == "":
            continue
        out[k] = v
        en = ZH_TO_EN.get(k)
        if en and en not in out:
            out[en] = v
    return out


def parse_xlsx(path: str):
    """解析一个数据 xlsx，返回 (metrics: dict[规范中文键]=数值, warnings: list[str])。"""
    metrics = {}
    warnings = []
    if not os.path.isfile(path):
        return metrics, [f"文件不存在：{path}"]
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        return metrics, [f"无法读取 Excel（{os.path.basename(path)}）：{e}"]

    ws = wb.active
    rows = []
    try:
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass

    max_r = len(rows)
    max_c = max((len(r) for r in rows), default=0)

    def cell(r, c):
        if 0 <= r < max_r and 0 <= c < len(rows[r]):
            return rows[r][c]
        return None

    for r in range(max_r):
        for c in range(max_c):
            v = cell(r, c)
            if v is None or v == "":
                continue
            key = _match_label(v)
            if not key:
                continue
            # 值优先取右邻格，其次下邻格，再斜下方
            val = None
            for (dr, dc) in ((0, 1), (1, 0), (1, 1), (0, 2)):
                cand = cell(r + dr, c + dc)
                if _is_value_like(cand) and _match_label(cand) is None:
                    val = cand
                    break
            if val is None:
                continue
            num = _to_value(val, key)
            if num is None:
                continue
            # 同一指标若已存在则保留第一个（避免把表头行错当数值覆盖）
            metrics.setdefault(key, num)

    if not metrics:
        warnings.append(f"未在「{os.path.basename(path)}」中识别到已知指标，请确认列名含：播放量/完播率/2秒跳出率/5秒完播率/平均播放时长/平均播放占比/涨粉量/涨粉率等")
    return metrics, warnings


def read_txt(path: str) -> str:
    """读取字幕稿 txt，返回去空白后的文本（失败返回空串）。兼容 utf-8 / gbk 编码。"""
    if not os.path.isfile(path):
        return ""
    raw = None
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            with open(path, "r", encoding=enc) as f:
                raw = f.read()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception as e:  # noqa: BLE001
            print(f"[import] 读取字幕稿失败({enc}): {e}")
            return ""
    if raw is None:
        return ""
    return _clean_subtitle(raw)


_TIME_LINE = re.compile(r"^\s*\d{1,2}:\d{2}:\d{2}[,.]\d{1,3}\s*-->\s*\d{1,2}:\d{2}:\d{2}[,.]\d{1,3}\s*(.*)$")


def _clean_subtitle(raw: str) -> str:
    """轻整理字幕稿：保留 SRT 时间戳行（时间 + 文本），去掉纯序号行/空行。"""
    lines = []
    for ln in raw.splitlines():
        s = ln.strip()
        if not s:
            continue
        # 纯 SRT 序号行（纯数字，后面跟着时间戳行）→ 跳过
        if re.fullmatch(r"\d+", s):
            continue
        lines.append(s)
    return "\n".join(lines).strip()


def parse_retention(path: str) -> str:
    """从数据 xlsx 里提取留存曲线（我的留存率 + 同类作品对标留存率），返回格式化文本。

    兼容「时间 | 留存率 | 同类作品留存率」或「时间 | 留存率」的排布。
    找不到返回空串。
    """
    if not os.path.isfile(path):
        return ""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return ""
    result = []
    try:
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            header = None
            my_col = bench_col = time_col = None
            # 1) 找含「留存」的表头，区分我的/同类
            for r, row in enumerate(rows):
                for c, v in enumerate(row):
                    if v is None or not str(v).strip():
                        continue
                    s = _norm_label(str(v))
                    if "留存" in s:
                        header = r
                        if ("同类" in s) or ("对标" in s) or ("同类型" in s):
                            bench_col = c
                        else:
                            my_col = c
                if header is not None:
                    break
            if header is None:
                continue
            # 2) 找时间列（表头行里的「时间/秒」）
            hrow = rows[header]
            for c, v in enumerate(hrow):
                if v is None:
                    continue
                s = _norm_label(str(v))
                if s in ("时间", "秒", "时间点") or s.startswith("时间"):
                    time_col = c
                    break
            if time_col is None:
                base = my_col if my_col is not None else (bench_col if bench_col is not None else 1)
                time_col = max(0, base - 1)
            # 3) 逐行读时间 + 留存率
            my_pts, bench_pts = [], []
            for r in range(header + 1, len(rows)):
                row = rows[r]
                t = row[time_col] if time_col < len(row) else None
                if t is None or not str(t).strip():
                    continue
                ts = str(t).strip()
                if not re.match(r"^\d+(?:\.\d+)?\s*(?:秒|s|S)?$", ts):
                    continue
                rate = row[my_col] if my_col is not None and my_col < len(row) else None
                bench = row[bench_col] if bench_col is not None and bench_col < len(row) else None
                rv = _to_value(rate, "留存率")
                bv = _to_value(bench, "留存率")
                if rv is not None:
                    my_pts.append(f"{ts} {rv}%")
                if bv is not None:
                    bench_pts.append(f"{ts} {bv}%")
            if my_pts or bench_pts:
                if my_pts:
                    result.append("我的留存曲线（每秒留存率）：" + "，".join(my_pts))
                if bench_pts:
                    result.append("同类作品留存曲线（对标）：" + "，".join(bench_pts))
                break
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass
    return "\n".join(result)
